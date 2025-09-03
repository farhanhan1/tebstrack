from flask import Blueprint, render_template, redirect, url_for, request, flash, session, current_app, jsonify, send_from_directory, make_response
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField
from wtforms.validators import DataRequired, Length
from flask_login import login_user, login_required, logout_user, current_user
from app.models import User, Ticket, db, Category, SystemSettings, EmailTemplate, TemplateActionStep, TicketTemplateRecommendation
from app.models import Log as TicketLog
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from sqlalchemy import func
import datetime
import csv
import io
from .fetch_emails_util import fetch_and_store_emails
import logging
from app.extensions import csrf
import bleach
from app.ai_service import get_ai_service
import os
main = Blueprint('main', __name__)

@main.route('/reply_ticket/<int:ticket_id>', methods=['POST'])
@login_required
@csrf.exempt
def reply_ticket(ticket_id):
    from .models import Ticket, EmailMessage, Log
    import bleach, datetime, uuid, smtplib, os
    from email.message import EmailMessage as PyEmailMessage
    import logging
    ticket = Ticket.query.get_or_404(ticket_id)
    body = bleach.clean(request.form.get('reply_body', ''))
    if not body:
        flash('Reply cannot be empty.', 'error')
        return redirect(url_for('main.view_ticket', ticket_id=ticket_id))
    # Compose and send the email
    smtp_host = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
    smtp_port = int(os.environ.get('SMTP_PORT', 587))
    smtp_user = os.environ.get('GMAIL_USER', 'tebstrack@gmail.com')
    smtp_pass = os.environ.get('GMAIL_APP_PASSWORD', '')
    import email.utils
    # Extract just the email address for sending
    _, to_addr = email.utils.parseaddr(ticket.sender)
    subject = f"Re: {ticket.subject}"
    # Debug: print SMTP config and reply info
    print(f"[DEBUG] SMTP config: host={smtp_host}, port={smtp_port}, user={smtp_user}", flush=True)
    print(f"[DEBUG] Sending reply to: {to_addr}, subject: {subject}", flush=True)
    logging.warning(f"[DEBUG] SMTP config: host={smtp_host}, port={smtp_port}, user={smtp_user}")
    logging.warning(f"[DEBUG] Sending reply to: {to_addr}, subject: {subject}")
    # Generate a unique message_id for this reply
    message_id = f"<tebstrack-{uuid.uuid4()}@tebstrack>"
    # Prevent duplicate by message_id
    if EmailMessage.query.filter_by(message_id=message_id).first():
        flash('Duplicate reply detected.', 'error')
        return redirect(url_for('main.view_ticket', ticket_id=ticket_id))
    # Build the email
    msg = PyEmailMessage()
    msg['Subject'] = subject
    msg['From'] = smtp_user
    msg['To'] = to_addr
    msg['Message-ID'] = message_id
    msg.set_content(body)
    # Optionally, set In-Reply-To and References headers if available
    if ticket.thread_id:
        last_msg = EmailMessage.query.filter_by(thread_id=ticket.thread_id).order_by(EmailMessage.sent_at.desc()).first()
        if last_msg and last_msg.message_id:
            msg['In-Reply-To'] = last_msg.message_id
            msg['References'] = last_msg.message_id
    email_sent = False
    email_error = None
    # Always extract the correct email address from ticket.sender
    import email.utils
    _, to_addr = email.utils.parseaddr(ticket.sender)
    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
        email_sent = True
        print("[DEBUG] Email sent successfully.", flush=True)
        logging.warning("[DEBUG] Email sent successfully.")
    except Exception as e:
        email_error = str(e)
        print(f"[DEBUG] Failed to send email: {e}", flush=True)
        logging.error(f"[DEBUG] Failed to send email: {e}")
        flash(f'Failed to send email: {e}', 'error')

    # Only save EmailMessage if email was sent successfully
    if email_sent:
        thread_id = ticket.thread_id or str(ticket.id)
        try:
            email_msg = EmailMessage(
                ticket_id=ticket.id,
                thread_id=thread_id,
                sender=current_user.username,
                subject=subject,
                body=body,
                sent_at=datetime.datetime.now(),
                attachments=None,
                message_id=message_id,
                in_reply_to=msg['In-Reply-To'] if 'In-Reply-To' in msg else None
            )
            db.session.add(email_msg)
            db.session.commit()
            log = Log(user=current_user.username, action='reply_ticket', details=f"Replied to ticket '{ticket.subject}' (ID: {ticket.id}) and sent email to {to_addr}")
            db.session.add(log)
            db.session.commit()
            flash('Reply sent as email and saved to thread.', 'success')
        except Exception as e:
            db.session.rollback()
            print(f"[DEBUG] Failed to save reply to DB: {e}", flush=True)
            logging.error(f"[DEBUG] Failed to save reply to DB: {e}")
            flash(f'Failed to save reply to thread: {e}', 'error')
    # Always refresh the page to show the updated thread, even if there was an error
    return redirect(url_for('main.view_ticket', ticket_id=ticket_id))


# Flask-WTF LoginForm for CSRF and validation
class LoginForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=1, max=64)])
    password = PasswordField('Password', validators=[DataRequired(), Length(min=1, max=128)])

# Favicon route (must be after Blueprint definition)
@main.route('/favicon.ico')
def favicon():
    return send_from_directory(current_app.root_path + '/../lib', 'tebstrack.ico', mimetype='image/vnd.microsoft.icon')

@main.route('/settings', methods=['GET'])
@login_required
def settings():
    from app.models import UserSettings, SystemSettings
    # Get user settings for all users
    user_settings = UserSettings.get_user_settings(current_user.id)
    
    # Only admin can manage categories and system settings
    if current_user.role != 'admin':
        all_categories = []
        system_settings = {}
    else:
        all_categories = Category.query.order_by(Category.name).all()
        # Get system settings for admin
        env_key = os.getenv('OPENAI_API_KEY', '')
        custom_key = SystemSettings.get_setting('openai_api_key', '')
        system_settings = {
            'openai_api_key': custom_key,
            'using_env_key': not bool(custom_key),
            'env_api_key': env_key
        }
    
    return render_template('settings.html', 
                         all_categories=all_categories, 
                         current_user=current_user,
                         user_settings=user_settings,
                         system_settings=system_settings)

@main.route('/update_pagination_settings', methods=['POST'])
@login_required
def update_pagination_settings():
    from app.models import UserSettings
    
    pagination_enabled = request.form.get('pagination_enabled') == 'on'
    tickets_per_page = request.form.get('tickets_per_page', 10, type=int)
    
    # Validate tickets_per_page - allow custom values from 1 to 500
    if tickets_per_page < 1:
        tickets_per_page = 1
        flash('Tickets per page cannot be less than 1. Set to minimum value of 1.', 'warning')
    elif tickets_per_page > 500:
        tickets_per_page = 500
        flash('Tickets per page cannot exceed 500. Set to maximum value of 500.', 'warning')
    
    # Get or create user settings
    user_settings = UserSettings.get_user_settings(current_user.id)
    user_settings.pagination_enabled = pagination_enabled
    user_settings.tickets_per_page = tickets_per_page
    
    db.session.commit()
    
    if pagination_enabled:
        flash(f'Pagination enabled with {tickets_per_page} tickets per page!', 'success')
    else:
        flash('Pagination disabled successfully!', 'success')
    return redirect(url_for('main.settings'))

@main.route('/update_openai_api_key', methods=['POST'])
@login_required 
def update_openai_api_key():
    from app.models import SystemSettings, Log
    from app.ai_service import reset_ai_service
    
    # Only admin can update system settings
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('main.settings'))
    
    try:
        api_key = request.form.get('openai_api_key', '').strip()
        use_env_key = request.form.get('use_env_key') == 'on'
        
        if use_env_key:
            # Remove custom API key to fall back to environment variable
            setting = SystemSettings.query.filter_by(setting_key='openai_api_key').first()
            if setting:
                db.session.delete(setting)
                db.session.commit()
            
            # Reset AI service to pick up new settings
            reset_ai_service()
            
            # Log the action
            log = Log(
                user=current_user.username,
                action='OpenAI API Key Updated',
                details='Switched to using environment variable for OpenAI API key'
            )
            db.session.add(log)
            db.session.commit()
            
            flash('Successfully switched to using environment variable for OpenAI API key!', 'success')
        else:
            if not api_key:
                flash('Please provide an API key or select "Use Environment Variable".', 'error')
                return redirect(url_for('main.settings'))
            
            # Basic validation - OpenAI keys typically start with 'sk-'
            if not api_key.startswith('sk-'):
                flash('Invalid API key format. OpenAI API keys typically start with "sk-".', 'error')
                return redirect(url_for('main.settings'))
            
            # Store custom API key
            SystemSettings.set_setting(
                'openai_api_key', 
                api_key,
                'Custom OpenAI API key for TeBSTrack AI services'
            )
            
            # Reset AI service to pick up new API key
            reset_ai_service()
            
            # Log the action (mask the key in logs)
            masked_key = api_key[:8] + '*' * (len(api_key) - 12) + api_key[-4:] if len(api_key) > 12 else 'sk-***'
            log = Log(
                user=current_user.username,
                action='OpenAI API Key Updated', 
                details=f'Updated OpenAI API key (ending in ...{api_key[-4:] if len(api_key) >= 4 else "***"})'
            )
            db.session.add(log)
            db.session.commit()
            
            flash('OpenAI API key updated successfully!', 'success')
        
        return redirect(url_for('main.settings'))
        
    except Exception as e:
        flash(f'Error updating API key: {str(e)}', 'error')
        return redirect(url_for('main.settings'))

@main.route('/add_category', methods=['POST'])
@login_required
def add_category():
    if current_user.role != 'admin':
        return redirect(url_for('main.settings'))
    new_category = request.form.get('category')
    if new_category:
        if not Category.query.filter_by(name=new_category).first():
            db.session.add(Category(name=new_category))
            db.session.commit()
            log = TicketLog(user=current_user.username, action='add_category', details=f"Added category '{new_category}'")
            db.session.add(log)
            db.session.commit()
    return redirect(url_for('main.settings'))

@main.route('/edit_category/<category>', methods=['POST'])
@login_required
def edit_category(category):
    if current_user.role != 'admin':
        return redirect(url_for('main.settings'))
    new_category = request.form.get('new_category')
    if new_category:
        cat = Category.query.filter_by(name=category).first()
        if cat:
            old_name = cat.name
            cat.name = new_category
            db.session.commit()
            # Update all tickets with the old category to the new one
            Ticket.query.filter_by(category=category).update({'category': new_category})
            db.session.commit()
            log = TicketLog(user=current_user.username, action='edit_category', details=f"Renamed category '{old_name}' to '{new_category}'")
            db.session.add(log)
            db.session.commit()
    return redirect(url_for('main.settings'))

@main.route('/delete_category/<category>', methods=['POST'])
@login_required
def delete_category(category):
    if current_user.role != 'admin':
        return redirect(url_for('main.settings'))
    cat = Category.query.filter_by(name=category).first()
    if cat:
        db.session.delete(cat)
        db.session.commit()
        # Remove category from all tickets
        Ticket.query.filter_by(category=category).update({'category': None})
        db.session.commit()
        log = TicketLog(user=current_user.username, action='delete_category', details=f"Deleted category '{category}'")
        db.session.add(log)
        db.session.commit()
    return redirect(url_for('main.settings'))



# Admin: Manage Users page
@main.route('/manage_users', methods=['GET'])
@login_required
def manage_users():
    if current_user.role != 'admin':
        return redirect(url_for('main.index'))
    users = User.query.order_by(User.username).all()
    return render_template('manage_users.html', users=users)

# Admin: Create User endpoint
@main.route('/create_user', methods=['POST'])
@login_required
def create_user():
    if current_user.role != 'admin':
        return redirect(url_for('main.index'))
    username = request.form.get('username')
    password = request.form.get('password')
    role = request.form.get('role')
    if not username or not password or not role:
        flash('All fields are required.', 'error')
        return redirect(url_for('main.manage_users'))
    if User.query.filter_by(username=username).first():
        flash('Username already exists.', 'error')
        return redirect(url_for('main.manage_users'))
    new_user = User(
        username=username,
        password=generate_password_hash(password),
        role=role
    )
    db.session.add(new_user)
    db.session.commit()
    from .models import Log
    log = Log(user=current_user.username, action='create_user', details=f"Created user '{username}' with role '{role}'")
    db.session.add(log)
    db.session.commit()
    flash(f"User '{username}' created.", 'success')
    return redirect(url_for('main.manage_users'))

# Admin: Delete User endpoint
@main.route('/delete_user/<username>', methods=['POST'])
@login_required
def delete_user(username):
    if current_user.role != 'admin':
        return redirect(url_for('main.index'))
    user = User.query.filter_by(username=username).first()
    if not user or user.username == 'admin':
        flash('Cannot delete this user.', 'error')
        return redirect(url_for('main.manage_users'))
    db.session.delete(user)
    db.session.commit()
    from .models import Log
    log = Log(user=current_user.username, action='delete_user', details=f"Deleted user '{username}'")
    db.session.add(log)
    db.session.commit()
    flash(f"User '{username}' deleted.", 'success')
    return redirect(url_for('main.manage_users'))

# Admin: Edit User endpoint
@main.route('/edit_user/<username>', methods=['POST'])
@login_required
def edit_user(username):
    if current_user.role != 'admin':
        return redirect(url_for('main.index'))
    user = User.query.filter_by(username=username).first()
    if not user or user.username == 'admin':
        flash('Cannot edit this user.', 'error')
        return redirect(url_for('main.manage_users'))
    password = request.form.get('password')
    role = request.form.get('role')
    if role:
        user.role = role
    if password:
        from werkzeug.security import generate_password_hash
        user.password = generate_password_hash(password)
    db.session.commit()
    from .models import Log
    log = Log(user=current_user.username, action='edit_user', details=f"Edited user '{username}' (role: {role})")
    db.session.add(log)
    db.session.commit()
    flash(f"User '{username}' updated.", 'success')
    return redirect(url_for('main.manage_users'))

# Admin: Audit Logs page
@main.route('/audit_logs', methods=['GET'])
@login_required
def audit_logs():
    if current_user.role != 'admin':
        return redirect(url_for('main.index'))
    from .models import Log
    logs = Log.query.order_by(Log.timestamp.desc()).all()
    return render_template('audit_logs.html', logs=logs)

# Create Ticket endpoint (must be after Blueprint definition)
@main.route('/create_ticket', methods=['POST'])
@login_required
@csrf.exempt
def create_ticket():
    # Sanitize all user input fields
    subject = bleach.clean(request.form.get('subject', ''))
    category = bleach.clean(request.form.get('category', ''))
    urgency = bleach.clean(request.form.get('urgency', ''))
    description = bleach.clean(request.form.get('description', ''))
    if not subject or not category or not urgency or not description:
        return jsonify({'success': False, 'error': 'All fields are required.'})
    try:
        # Try to use current_user.email if available, else just username
        sender = None
        if hasattr(current_user, 'email') and current_user.email:
            sender = f"{current_user.username} <{current_user.email}>"
        else:
            sender = current_user.username if hasattr(current_user, 'username') else 'Unknown'
        print(f"[DEBUG] Manual ticket sender: {sender}", flush=True)
        ticket = Ticket(
            subject=subject,
            category=category,
            urgency=urgency,
            description=description,
            sender=sender,  # Don't bleach the sender field to preserve email format
            status='Open',
            created_at=datetime.datetime.now()
        )
        db.session.add(ticket)
        db.session.commit()
        from .models import Log
        log = Log(user=current_user.username, action='create_ticket', details=f"Created ticket '{ticket.subject}' (ID: {ticket.id})")
        db.session.add(log)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})
@main.route('/bulk_ticket_action', methods=['POST'])
@login_required
def bulk_ticket_action():
    data = request.get_json()
    action = data.get('action')
    ticket_ids = data.get('ticket_ids', [])
    if not ticket_ids or not action:
        return jsonify({'success': False, 'error': 'Missing ticket IDs or action.'}), 400

    tickets = Ticket.query.filter(Ticket.id.in_(ticket_ids)).all()
    
    # Role-based access control - filter tickets user can modify
    if current_user.role == 'infra':
        # Infra users can only modify unassigned tickets or their own assigned tickets
        allowed_tickets = []
        for ticket in tickets:
            if ticket.assigned_to is None or ticket.assigned_to == current_user.id:
                allowed_tickets.append(ticket)
        tickets = allowed_tickets
        
        # Infra users cannot delete tickets
        if action == 'delete':
            return jsonify({'success': False, 'error': 'You do not have permission to delete tickets.'}), 403
    elif current_user.role != 'admin':
        # Only admin and infra can perform bulk actions
        return jsonify({'success': False, 'error': 'You do not have permission to perform bulk actions.'}), 403

    affected = 0
    username = current_user.username if hasattr(current_user, 'username') else 'Unknown'
    debug_info = {}
    from app.models import EmailMessage
    deleted_email_count = 0
    for ticket in tickets:
        if action == 'close' and ticket.status != 'Closed':
            ticket.status = 'Closed'
            db.session.add(TicketLog(user=username, action='Bulk Close', details=f'Ticket closed in bulk for ticket {ticket.id}'))
            affected += 1
        elif action == 'open' and ticket.status != 'Open':
            ticket.status = 'Open'
            db.session.add(TicketLog(user=username, action='Bulk Open', details=f'Ticket opened in bulk for ticket {ticket.id}'))
            affected += 1
        elif action == 'delete' and current_user.role == 'admin':
            # Only admin can delete
            count = EmailMessage.query.filter(EmailMessage.ticket_id == ticket.id).delete(synchronize_session=False)
            deleted_email_count += count
            db.session.add(TicketLog(user=username, action='Bulk Delete', details=f'Ticket deleted in bulk for ticket {ticket.id} (deleted {count} related emails)'))
            db.session.delete(ticket)
            affected += 1
        # Future actions can be added here
    if action == 'delete':
        debug_info['deleted_email_count'] = deleted_email_count
        debug_info['deleted_ticket_count'] = affected
        print(f"[DEBUG] Bulk delete: {debug_info}", flush=True)
        logging.warning(f"[DEBUG] Bulk delete: {debug_info}")
    try:
        db.session.commit()
        return jsonify({'success': True, 'affected': affected})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    try:
        ticket = Ticket(
            subject=subject,
            category=category,
            urgency=urgency,
            description=description,
            sender=current_user.username if hasattr(current_user, 'username') else 'Unknown',
            status='Open',
            created_at=datetime.datetime.now()
        )
        db.session.add(ticket)
        db.session.commit()
        from .models import Log
        log = Log(user=current_user.username, action='create_ticket', details=f"Created ticket '{ticket.subject}' (ID: {ticket.id})")
        db.session.add(log)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})



# Fetch emails endpoint (must be after main is defined)

@main.route('/fetch_emails', methods=['POST'])
@login_required
@csrf.exempt
def fetch_emails():
    try:
        count = fetch_and_store_emails()
        return jsonify({'success': True, 'new_tickets': count}), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 200



# Delete tickets route (must be after Blueprint definition, and only defined once)
@main.route('/delete_tickets', methods=['POST'])
@login_required
def delete_tickets():
    # Only admin users can delete tickets
    if current_user.role != 'admin':
        if request.is_json:
            return jsonify({'success': False, 'error': 'You do not have permission to delete tickets.'}), 403
        flash('You do not have permission to delete tickets.', 'error')
        return redirect(url_for('main.tickets'))
    
    ids = request.form.getlist('ticket_ids')
    if not ids and request.is_json:
        ids = request.get_json().get('ticket_ids', [])
    if ids:
        from app.models import EmailMessage, Ticket
        # Convert ids to integers to avoid type mismatch
        ticket_ids = [int(i) for i in ids]
        # Debug info
        debug_info = {}
        debug_info['ticket_ids'] = ticket_ids
        # Delete all EmailMessages with ticket_id in ticket_ids
        deleted_emails = EmailMessage.query.filter(EmailMessage.ticket_id.in_(ticket_ids)).delete(synchronize_session=False)
        debug_info['deleted_emails'] = deleted_emails
        # Now delete the tickets
        deleted_tickets = Ticket.query.filter(Ticket.id.in_(ticket_ids)).delete(synchronize_session=False)
        debug_info['deleted_tickets'] = deleted_tickets
        # Use both print and logging to ensure output appears in VS Code
        print(f"[DEBUG] Ticket deletion: {debug_info}", flush=True)
        logging.warning(f"[DEBUG] Ticket deletion: {debug_info}")
        # If you still do not see output, run Flask with unbuffered output:
        #   python -u run.py
        db.session.commit()
        if request.is_json:
            return jsonify({'success': True, 'debug': debug_info})
    if request.is_json:
        return jsonify({'success': False, 'error': 'No ticket IDs provided.'})
    return redirect(url_for('main.tickets'))

# (Removed duplicate import and Blueprint block)

@main.before_app_request
def require_login():
    allowed_routes = ['main.login', 'static']
    if not current_user.is_authenticated and request.endpoint not in allowed_routes:
        return redirect(url_for('main.login'))

@main.route('/')
def index():
    # Quick stats - filtered by role
    if current_user.role == 'infra':
        from sqlalchemy import or_
        # Infra users stats only for tickets they can see
        open_count = Ticket.query.filter_by(status='Open').filter(or_(
            Ticket.assigned_to == None,
            Ticket.assigned_to == current_user.id
        )).count()
        urgent_count = Ticket.query.filter_by(urgency='Urgent', status='Open').filter(or_(
            Ticket.assigned_to == None,
            Ticket.assigned_to == current_user.id
        )).count()
        closed_count = Ticket.query.filter_by(status='Closed').filter(or_(
            Ticket.assigned_to == None,
            Ticket.assigned_to == current_user.id
        )).count()
    else:
        # Admin users see all tickets
        open_count = Ticket.query.filter_by(status='Open').count()
        urgent_count = Ticket.query.filter_by(urgency='Urgent', status='Open').count()
        closed_count = Ticket.query.filter_by(status='Closed').count()
    
    user_count = User.query.count()

    # Top requestor (by sender) - filtered by role
    if current_user.role == 'infra':
        from sqlalchemy import or_
        top = db.session.query(Ticket.sender, func.count(Ticket.id).label('count')) \
            .filter(or_(
                Ticket.assigned_to == None,
                Ticket.assigned_to == current_user.id
            )) \
            .group_by(Ticket.sender) \
            .order_by(func.count(Ticket.id).desc()) \
            .first()
    else:
        top = db.session.query(Ticket.sender, func.count(Ticket.id).label('count')) \
            .group_by(Ticket.sender) \
            .order_by(func.count(Ticket.id).desc()) \
            .first()
    top_requestor = {'name': top[0], 'count': top[1]} if top else None

    # Ticket stats by month (last 6 months)
    months_to_show = 6
    now = datetime.datetime.now()
    months = []
    open_counts = []
    closed_counts = []
    for i in range(months_to_show-1, -1, -1):
        month = (now - datetime.timedelta(days=now.day-1)).replace(day=1) - datetime.timedelta(days=30*i)
        month_start = month.replace(day=1)
        next_month = (month_start + datetime.timedelta(days=32)).replace(day=1)
        label = month_start.strftime('%Y-%m')
        months.append(label)
        open_count_month = Ticket.query.filter(Ticket.status=='Open', Ticket.created_at >= month_start, Ticket.created_at < next_month).count()
        closed_count_month = Ticket.query.filter(Ticket.status=='Closed', Ticket.created_at >= month_start, Ticket.created_at < next_month).count()
        open_counts.append(open_count_month)
        closed_counts.append(closed_count_month)
    ticket_stats_by_month = {
        'months': months,
        'open': open_counts,
        'closed': closed_counts
    }

    # Tickets list (for table) - filtered by role
    if current_user.role == 'infra':
        from sqlalchemy import or_
        tickets = Ticket.query.filter(or_(
            Ticket.assigned_to == None,
            Ticket.assigned_to == current_user.id
        )).order_by(Ticket.created_at.desc()).all()
    else:
        tickets = Ticket.query.order_by(Ticket.created_at.desc()).all()

    return render_template('home.html',
        open_count=open_count,
        urgent_count=urgent_count,
        closed_count=closed_count,
        user_count=user_count,
        tickets=tickets,
        current_user=current_user,
        ticket_stats_by_month=ticket_stats_by_month,
        months_to_show=months_to_show,
        top_requestor=top_requestor
    )

@main.route('/login', methods=['GET', 'POST'])
def login():
    import datetime
    import logging
    import time
    form = LoginForm()
    # Brute force protection: server-side using LoginAttempt model
    from app.models import LoginAttempt
    max_attempts = 5
    lockout_minutes = 15
    ip = request.remote_addr
    now_ts = time.time()
    attempt = LoginAttempt.query.filter_by(ip=ip).first()
    from app.models import Log
    if attempt and attempt.lockout_until and now_ts < attempt.lockout_until:
        flash('Too many failed login attempts. Try again later.', 'error')
        logging.warning(f"Locked out login attempt from {ip}")
        # Log to DB
        log = Log(user=ip, action='login_lockout', details=f"Locked out login attempt from {ip}")
        db.session.add(log)
        db.session.commit()
        return render_template('login.html', form=form)
    # Reset fail count if lockout expired
    if attempt and attempt.lockout_until and now_ts >= attempt.lockout_until:
        attempt.fail_count = 0
        attempt.lockout_until = 0
        from app import db
        db.session.commit()
    if form.validate_on_submit():
        username = form.username.data.strip()
        password = form.password.data
        # Backend input validation: reject empty or whitespace-only username/password
        if not username or not password or username.isspace() or password.isspace():
            flash('Invalid username or password.', 'error')
            logging.warning(f"Login input validation failed from {ip} (username: '{username}')")
            log = Log(user=ip, action='login_input_invalid', details=f"Input validation failed (username: '{username}')")
            db.session.add(log)
            db.session.commit()
            return render_template('login.html', form=form)
        # Optionally: add regex for allowed characters here
        user = User.query.filter_by(username=username).first()
        from app import db
        if user and check_password_hash(user.password, password):
            login_user(user)
            # Reset fail count on successful login
            if attempt:
                attempt.fail_count = 0
                attempt.lockout_until = 0
                db.session.commit()
            logging.info(f"Login success for {username} from {ip}")
            log = Log(user=username, action='login_success', details=f"Login success from {ip}")
            db.session.add(log)
            db.session.commit()
            return redirect(url_for('main.index'))
        else:
            # Brute force: increment fail count in DB
            if not attempt:
                attempt = LoginAttempt(ip=ip, fail_count=1, lockout_until=0)
                db.session.add(attempt)
            else:
                attempt.fail_count += 1
            attempts_left = max_attempts - attempt.fail_count
            if attempt.fail_count >= max_attempts:
                attempt.lockout_until = now_ts + (lockout_minutes * 60)
                db.session.commit()
                flash('Too many failed login attempts. Try again later.', 'error')
                logging.warning(f"Account lockout for {username} from {ip}")
                log = Log(user=username, action='login_lockout', details=f"Account lockout from {ip}")
                db.session.add(log)
                db.session.commit()
            else:
                db.session.commit()
                flash(f'Login failed. {attempts_left} attempt(s) left before lockout.', 'error')
                logging.warning(f"Login failed for {username} from {ip} (attempt {attempt.fail_count})")
                log = Log(user=username, action='login_failed', details=f"Login failed from {ip} (attempt {attempt.fail_count})")
                db.session.add(log)
                db.session.commit()
            return render_template('login.html', form=form)
    # Always pass form to template
    return render_template('login.html', form=form)

@main.route('/logout')
@login_required
def logout():
    from app.models import Log
    username = current_user.username if hasattr(current_user, 'username') else 'Unknown'
    log = Log(user=username, action='logout', details=f"User {username} logged out.")
    db.session.add(log)
    db.session.commit()
    logout_user()
    flash('You have successfully logged out!', 'info')
    return redirect(url_for('main.login'))

@main.route('/profile')
@login_required
def profile():
    return render_template('profile.html')

@main.route('/test-session')
def test_session():
    session['test'] = 'hello'
    return f"Session set to: {session['test']}"

@main.route('/tickets')
@login_required
def tickets():
    from app.models import UserSettings
    
    # Get user pagination settings
    user_settings = UserSettings.get_user_settings(current_user.id)
    
    # Filters
    month = request.args.get('month', 'All')
    status = request.args.get('status', 'All')
    category = request.args.get('category')
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    
    import datetime
    query = Ticket.query
    
    # Role-based filtering: infra users can only see unassigned tickets and their own assigned tickets
    if current_user.role == 'infra':
        from sqlalchemy import or_
        query = query.filter(or_(
            Ticket.assigned_to == None,  # Unassigned tickets
            Ticket.assigned_to == current_user.id  # Their own assigned tickets
        ))
    # Admin users can see all tickets (no additional filtering needed)
    
    if month != 'All':
        try:
            month_start = datetime.datetime.strptime(month, '%Y-%m')
            next_month = (month_start + datetime.timedelta(days=32)).replace(day=1)
            query = query.filter(Ticket.created_at >= month_start, Ticket.created_at < next_month)
        except Exception:
            pass  # fallback: show all if parsing fails
    if status and status != 'All':
        query = query.filter(Ticket.status == status)
    if category and category != 'All':
        query = query.filter(Ticket.category == category)
    
    # Apply pagination if enabled
    if user_settings.pagination_enabled:
        pagination = query.order_by(Ticket.created_at.desc()).paginate(
            page=page,
            per_page=user_settings.tickets_per_page,
            error_out=False
        )
        tickets = pagination.items
        total_tickets = pagination.total
    else:
        tickets = query.order_by(Ticket.created_at.desc()).all()
        pagination = None
        total_tickets = len(tickets)

    # Key stats (use all tickets for stats, not just current page)
    all_tickets_query = query
    if user_settings.pagination_enabled:
        all_tickets_for_stats = all_tickets_query.all()
        tickets_raised = len(all_tickets_for_stats)
    else:
        tickets_raised = len(tickets)
        all_tickets_for_stats = tickets
    
    from collections import Counter
    most_common_category = Counter([t.category for t in all_tickets_for_stats if t.category]).most_common(1)
    most_common_category = most_common_category[0][0] if most_common_category else '-'
    most_common_requestor = Counter([t.sender for t in all_tickets_for_stats if t.sender]).most_common(1)
    most_common_requestor = most_common_requestor[0][0] if most_common_requestor else '-'

    # For dropdowns
    # Get all months for dropdown (YYYY-MM format) from tickets in DB
    from sqlalchemy import func
    all_months = [m[0] for m in db.session.query(func.strftime('%Y-%m', Ticket.created_at)).distinct().order_by(func.strftime('%Y-%m', Ticket.created_at).desc()).all()]
    all_categories = [c.name for c in Category.query.order_by(Category.name).all()]
    all_statuses = ['Open', 'Closed', 'All']

    user_map = {u.id: u.username for u in User.query.all()}
    return render_template('tickets.html',
        tickets=tickets,
        month=month,
        status=status,
        category=category,
        all_months=all_months,
        all_categories=all_categories,
        all_statuses=all_statuses,
        tickets_raised=tickets_raised,
        most_common_category=most_common_category,
        most_common_requestor=most_common_requestor,
        user_map=user_map,
        pagination=pagination,
        user_settings=user_settings
    )

@main.route('/export_tickets', methods=['GET', 'POST'])
@login_required
def export_tickets():
    """Export tickets to CSV or XLSX format"""
    export_format = request.args.get('format', 'csv').lower()
    month = request.args.get('month', 'All')
    status = request.args.get('status', 'All')
    category = request.args.get('category', 'All')
    
    # Handle None or empty string values
    if not month or month == 'None':
        month = 'All'
    if not status or status == 'None':
        status = 'All'
    if not category or category == 'None':
        category = 'All'
    
    # Debug info
    print(f"[DEBUG] Export params: format={export_format}, month={month}, status={status}, category={category}", flush=True)
    
    # Check if specific ticket IDs were provided (for selected tickets)
    selected_ticket_ids = None
    if request.method == 'POST':
        data = request.get_json()
        if data and 'ticket_ids' in data:
            selected_ticket_ids = data['ticket_ids']
            print(f"[DEBUG] Selected ticket IDs: {selected_ticket_ids}", flush=True)
    
    # Apply same filtering logic as tickets route
    query = Ticket.query
    
    # If specific ticket IDs provided, filter by those first
    if selected_ticket_ids:
        query = query.filter(Ticket.id.in_(selected_ticket_ids))
    
    # Role-based filtering
    if current_user.role == 'infra':
        from sqlalchemy import or_
        query = query.filter(or_(
            Ticket.assigned_to == None,
            Ticket.assigned_to == current_user.id
        ))
    
    # Apply filters only if not exporting selected tickets
    if not selected_ticket_ids:
        if month != 'All':
            try:
                year, month_num = month.split('-')
                start_date = datetime.datetime(int(year), int(month_num), 1)
                if int(month_num) == 12:
                    end_date = datetime.datetime(int(year) + 1, 1, 1)
                else:
                    end_date = datetime.datetime(int(year), int(month_num) + 1, 1)
                query = query.filter(Ticket.created_at >= start_date, Ticket.created_at < end_date)
            except ValueError:
                pass
        
        if status and status != 'All':
            query = query.filter(Ticket.status == status)
        
        if category and category != 'All':
            query = query.filter(Ticket.category == category)
    
    tickets = query.order_by(Ticket.created_at.desc()).all()
    print(f"[DEBUG] Found {len(tickets)} tickets for export", flush=True)
    
    user_map = {u.id: u.username for u in User.query.all()}
    
    # Prepare data for export
    headers = [
        'S/N', 'Issue Reported Date', 'Category', 'Ticket Name', 'Request by',
        'Status', 'Resolution', 'Assignee', 'Urgency', 'Date of Completion'
    ]
    
    rows = []
    for ticket in tickets:
        row = [
            ticket.id,
            ticket.created_at.strftime('%Y-%m-%d') if ticket.created_at else '-',
            ticket.category or '-',
            ticket.subject,
            ticket.sender if ticket.sender else '-',
            ticket.status,
            ticket.resolution if ticket.status == 'Closed' and ticket.resolution else '-',
            user_map.get(ticket.assigned_to, '-') if ticket.assigned_to else '-',
            ticket.urgency or '-',
            ticket.updated_at.strftime('%Y-%m-%d') if ticket.status == 'Closed' and ticket.updated_at else '-'
        ]
        rows.append(row)
    
    if export_format == 'xlsx':
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tickets"
            
            # Add headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add data
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].auto_size = True
            
            # Save to memory
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename=tickets_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            return response
            
        except ImportError:
            flash('XLSX export requires openpyxl package. Falling back to CSV.', 'warning')
            export_format = 'csv'
    
    if export_format == 'csv':
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=tickets_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        return response
    
    flash('Invalid export format specified.', 'error')
    return redirect(url_for('main.tickets'))


from flask_wtf import FlaskForm
from wtforms import StringField, TextAreaField, SelectField, IntegerField
from wtforms.validators import DataRequired, Optional

class EditTicketForm(FlaskForm):
    subject = StringField('Subject', validators=[DataRequired()])
    category = StringField('Category', validators=[DataRequired()])
    urgency = SelectField('Urgency', choices=[('Low','Low'),('Medium','Medium'),('High','High'),('Urgent','Urgent')], validators=[DataRequired()])
    status = SelectField('Status', choices=[('Open','Open'),('Closed','Closed')], validators=[DataRequired()])
    description = TextAreaField('Description', validators=[DataRequired()])
    assigned_to = IntegerField('Assignee', validators=[Optional()])
    resolution = StringField('Resolution', validators=[Optional()])
    sender = StringField('Request by', validators=[Optional()])
    created_at = StringField('Issue Reported Date', validators=[Optional()])  # Will parse as date
    updated_at = StringField('Date of Completion', validators=[Optional()])  # Will parse as date

@main.route('/viewticket/<int:ticket_id>')
@login_required
def view_ticket(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    
    # Role-based access control: infra users can only view unassigned tickets or their own assigned tickets
    if current_user.role == 'infra':
        if ticket.assigned_to is not None and ticket.assigned_to != current_user.id:
            flash('You do not have permission to view this ticket.', 'error')
            return redirect(url_for('main.tickets'))
    
    all_categories = Category.query.order_by(Category.name).all()
    from .models import User, EmailMessage
    
    # Filter users based on role for assignment dropdown
    if current_user.role == 'admin':
        all_users = User.query.order_by(User.username).all()
    else:
        # Infra users only see themselves in assignment options
        all_users = [current_user]
    # Fetch all EmailMessages for this ticket's thread, ordered by sent_at
    thread_msgs = []
    import os
    GMAIL_USER = os.environ.get('GMAIL_USER', 'tebstrack@gmail.com').lower()
    from sqlalchemy import or_, and_
    if ticket.thread_id:
        # Build the set of message_ids in the thread chain
        all_msgs = EmailMessage.query.filter(
            EmailMessage.thread_id == ticket.thread_id
        ).order_by(EmailMessage.sent_at.desc()).all()
        # Build a map of message_id -> EmailMessage
        msg_map = {m.message_id: m for m in all_msgs if m.message_id}
        # Find the root message (the one with in_reply_to=None or not in msg_map)
        root_msgs = [m for m in all_msgs if not m.in_reply_to or m.in_reply_to not in msg_map]
        thread_chain = []
        # For each root, walk the chain
        for root in root_msgs:
            chain = []
            current = root
            while current:
                chain.append(current)
                # Find the next message that replies to this one
                next_msg = None
                for m in all_msgs:
                    if m.in_reply_to and m.in_reply_to == current.message_id:
                        next_msg = m
                        break
                current = next_msg
            thread_chain.extend(chain)
        # Remove duplicates and sort by sent_at
        seen_ids = set()
        thread_msgs = []
        for m in sorted(thread_chain, key=lambda x: x.sent_at, reverse=True):
            if m.id not in seen_ids:
                thread_msgs.append(m)
                seen_ids.add(m.id)
    else:
        msg = EmailMessage.query.filter_by(ticket_id=ticket.id).first()
        if msg:
            thread_msgs = [msg]
    # Convert attachments JSON string to list for each message, and mark if sent by GMAIL_USER
    import re
    for msg in thread_msgs:
        try:
            import json
            msg.attachments = json.loads(msg.attachments) if msg.attachments else []
        except Exception:
            msg.attachments = []
        msg.is_self = (msg.sender or '').strip().lower() == GMAIL_USER
        # Remove quoted message history (e.g., lines starting with 'On ... wrote:')
        if msg.body:
            # Remove everything from the first occurrence of a quoted reply marker
            msg.body = re.split(r'\n?On .+wrote:', msg.body)[0].strip()
    form = EditTicketForm(obj=ticket)
    return render_template('viewticket.html', ticket=ticket, all_categories=all_categories, all_users=all_users, thread_msgs=thread_msgs, GMAIL_USER=GMAIL_USER, form=form)

# Route to serve attachments
import os
from flask import send_from_directory, abort
@main.route('/attachments/<path:filename>')
@login_required
def serve_attachment(filename):
    # attachments directory is at the project root
    # attachments directory is at the project root, not inside app/
    attachments_dir = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'attachments'))
    file_path = os.path.join(attachments_dir, filename)
    if not os.path.isfile(file_path):
        abort(404)
    return send_from_directory(attachments_dir, filename, as_attachment=True)

# Admin: Edit Ticket page

@main.route('/assign_ticket/<int:ticket_id>', methods=['POST'])
@login_required
def assign_ticket(ticket_id):
    """Allow infra users to assign/unassign themselves to/from tickets"""
    ticket = Ticket.query.get_or_404(ticket_id)
    
    # Role-based access control
    if current_user.role == 'infra':
        # Infra users can only view unassigned tickets or their own assigned tickets
        if ticket.assigned_to is not None and ticket.assigned_to != current_user.id:
            flash('You do not have permission to modify this ticket.', 'error')
            return redirect(url_for('main.tickets'))
        
        action = request.form.get('action')
        if action == 'assign':
            # Can only assign to themselves if ticket is unassigned
            if ticket.assigned_to is None:
                ticket.assigned_to = current_user.id
                flash('Ticket assigned to you successfully.', 'success')
                from .models import Log
                log = Log(user=current_user.username, action='assign_ticket', 
                         details=f"Self-assigned ticket '{ticket.subject}' (ID: {ticket.id})")
                db.session.add(log)
            else:
                flash('This ticket is already assigned.', 'error')
        elif action == 'unassign':
            # Can only unassign if assigned to themselves
            if ticket.assigned_to == current_user.id:
                ticket.assigned_to = None
                flash('Ticket unassigned successfully.', 'success')
                from .models import Log
                log = Log(user=current_user.username, action='unassign_ticket', 
                         details=f"Self-unassigned ticket '{ticket.subject}' (ID: {ticket.id})")
                db.session.add(log)
            else:
                flash('You can only unassign tickets assigned to you.', 'error')
        
        db.session.commit()
    else:
        flash('You do not have permission to perform this action.', 'error')
    
    return redirect(url_for('main.view_ticket', ticket_id=ticket_id))


@main.route('/edit_ticket/<int:ticket_id>', methods=['GET', 'POST'])
@login_required
def edit_ticket(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    
    # Role-based access control
    if current_user.role == 'infra':
        # Infra users can only edit tickets they can view (unassigned or their own)
        if ticket.assigned_to is not None and ticket.assigned_to != current_user.id:
            flash('You do not have permission to edit this ticket.', 'error')
            return redirect(url_for('main.tickets'))
    elif current_user.role != 'admin':
        # Only admin and infra users can edit tickets
        flash('You do not have permission to edit tickets.', 'error')
        return redirect(url_for('main.index'))
    
    infra_users = User.query.filter_by(role='infra').order_by(User.username).all()
    form = EditTicketForm(obj=ticket)
    
    if form.validate_on_submit():
        import bleach
        changes = []
        old = {
            'subject': ticket.subject,
            'category': ticket.category,
            'urgency': ticket.urgency,
            'status': ticket.status,
            'description': ticket.description,
            'assigned_to': ticket.assigned_to,
            'resolution': ticket.resolution,
            'sender': ticket.sender,
            'created_at': ticket.created_at.strftime('%Y-%m-%d') if ticket.created_at else '',
            'updated_at': ticket.updated_at.strftime('%Y-%m-%d') if ticket.updated_at else ''
        }
        new = {
            'subject': bleach.clean(form.subject.data),
            'category': bleach.clean(form.category.data),
            'urgency': bleach.clean(form.urgency.data),
            'status': bleach.clean(form.status.data),
            'description': bleach.clean(form.description.data),
            'assigned_to': form.assigned_to.data if form.assigned_to.data else None,
            'resolution': bleach.clean(form.resolution.data) if hasattr(form, 'resolution') and form.resolution.data else None,
            'sender': form.sender.data if hasattr(form, 'sender') and form.sender.data else ticket.sender,  # Don't bleach sender field
            'created_at': bleach.clean(form.created_at.data) if hasattr(form, 'created_at') and form.created_at.data else old['created_at'],
            'updated_at': bleach.clean(form.updated_at.data) if hasattr(form, 'updated_at') and form.updated_at.data else old['updated_at']
        }
        
        # Role-based assignment restrictions
        if current_user.role == 'infra':
            # Infra users can only assign to themselves or unassign
            if new['assigned_to'] is not None and new['assigned_to'] != current_user.id:
                flash('You can only assign tickets to yourself or leave them unassigned.', 'error')
                return render_template('edit_ticket.html', ticket=ticket, infra_users=infra_users, form=form)
            # Infra users cannot modify certain fields
            new['sender'] = old['sender']  # Keep original sender
            new['created_at'] = old['created_at']  # Keep original creation date
        
        user_map = {u.id: u.username for u in User.query.all()}
        for field in old:
            old_val = old[field]
            new_val = new[field]
            if field == 'assigned_to':
                old_disp = user_map.get(old_val, '-') if old_val else '-'
                new_disp = user_map.get(new_val, '-') if new_val else '-'
            else:
                old_disp = old_val if old_val is not None else '-'
                new_disp = new_val if new_val is not None else '-'
            if old_val != new_val:
                changes.append(f"{field} changed from '{old_disp}' to '{new_disp}'")
                if field in ['created_at', 'updated_at']:
                    # Parse date string to datetime
                    import datetime
                    try:
                        setattr(ticket, field, datetime.datetime.strptime(new_val, '%Y-%m-%d'))
                    except Exception:
                        pass  # Ignore parse errors
                else:
                    setattr(ticket, field, new_val)
        db.session.commit()
        from .models import Log
        if changes:
            details = f"Edited ticket '{ticket.subject}' (ID: {ticket.id}):\n" + "; ".join(changes)
        else:
            details = f"Edited ticket '{ticket.subject}' (ID: {ticket.id}): No changes."
        log = Log(user=current_user.username, action='edit_ticket', details=details)
        db.session.add(log)
        db.session.commit()
        return redirect(url_for('main.tickets'))
    # Pre-populate form with ticket data for GET
    if request.method == 'GET':
        form = EditTicketForm(obj=ticket)
    return render_template('edit_ticket.html', ticket=ticket, infra_users=infra_users, form=form)


# ============================================================================
# AI-POWERED ROUTES - OpenAI Integration for Enhanced Ticket Management
# ============================================================================

@main.route('/api/ai/categorize', methods=['POST'])
@login_required
@csrf.exempt
def ai_categorize_ticket():
    """AI-powered ticket categorization endpoint."""
    try:
        # Check if OpenAI API key is configured first
        if not os.getenv('OPENAI_API_KEY'):
            return jsonify({
                'success': False,
                'error': 'OpenAI API key not configured. Please set OPENAI_API_KEY in your environment variables.'
            }), 200  # Return 200 to avoid HTML error page
        
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'error': 'No data provided'
            }), 200
        
        subject = data.get('subject', '')
        body = data.get('body', '')
        sender = data.get('sender', '')
        
        if not subject and not body:
            return jsonify({
                'success': False,
                'error': 'Subject or body is required'
            }), 200
        
        ai_service = get_ai_service()
        result = ai_service.categorize_ticket(subject, body, sender)
        
        return jsonify({
            'success': True,
            'categorization': result
        })
        
    except Exception as e:
        logging.error(f"Error in AI categorization: {e}")
        return jsonify({
            'success': False,
            'error': f'Categorization failed: {str(e)}'
        }), 200


@main.route('/api/ai/recommend-template', methods=['POST'])
@login_required
@csrf.exempt
def ai_recommend_template():
    """AI-powered email template recommendation endpoint."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        subject = data.get('subject', '')
        body = data.get('body', '')
        category = data.get('category', '')
        
        if not subject and not body:
            return jsonify({'error': 'Subject or body is required'}), 400
        
        # Check if OpenAI API key is configured
        if not os.getenv('OPENAI_API_KEY'):
            return jsonify({'error': 'OpenAI API key not configured'}), 500
        
        ai_service = get_ai_service()
        result = ai_service.recommend_email_template(subject, body, category)
        
        return jsonify({
            'success': True,
            'templates': result
        })
        
    except Exception as e:
        logging.error(f"Error in template recommendation: {e}")
        return jsonify({'error': f'Template recommendation failed: {str(e)}'}), 500


@main.route('/api/templates/list', methods=['GET'])
@login_required
def get_all_templates():
    """API endpoint to get all email templates for dropdown selection."""
    try:
        templates = EmailTemplate.query.all()
        template_list = []
        
        for template in templates:
            template_list.append({
                'id': template.id,
                'name': template.name,
                'subject': template.subject,
                'use_case_description': template.use_case_description
            })
        
        return jsonify({
            'success': True,
            'templates': template_list
        })
        
    except Exception as e:
        logging.error(f"Error fetching templates: {e}")
        return jsonify({'error': f'Failed to fetch templates: {str(e)}'}), 500


@main.route('/api/templates/get/<template_name>', methods=['GET'])
@login_required
def get_template_by_name(template_name):
    """API endpoint to get a specific template by name."""
    try:
        template = EmailTemplate.query.filter_by(name=template_name).first()
        
        if not template:
            return jsonify({'error': 'Template not found'}), 404
        
        return jsonify({
            'success': True,
            'template': {
                'id': template.id,
                'name': template.name,
                'subject': template.subject,
                'body': template.body,
                'use_case_description': template.use_case_description
            }
        })
        
    except Exception as e:
        logging.error(f"Error fetching template: {e}")
        return jsonify({'error': f'Failed to fetch template: {str(e)}'}), 500


@main.route('/api/automation/vpn-creation/credentials', methods=['POST'])
@login_required
@csrf.exempt
def get_vpn_credentials():
    """Get VPN credentials with user confirmation popup"""
    try:
        data = request.get_json()
        sender_email = data.get('sender_email', '')
        
        if not sender_email:
            return jsonify({'error': 'Sender email is required'}), 400
        
        # Extract username from email
        from .automation_service import get_automation_service
        automation_service = get_automation_service()
        
        suggested_username = automation_service.extract_username_from_email(sender_email)
        generated_password = automation_service.generate_vpn_password()
        
        return jsonify({
            'success': True,
            'suggested_username': suggested_username,
            'generated_password': generated_password
        })
        
    except Exception as e:
        logging.error(f"Error generating VPN credentials: {e}")
        return jsonify({'error': f'Failed to generate credentials: {str(e)}'}), 500


@main.route('/api/automation/vpn-creation/execute', methods=['POST'])
@login_required
@csrf.exempt
def execute_vpn_creation():
    """Execute VPN account creation automation"""
    try:
        data = request.get_json()
        sender_email = data.get('sender_email', '')
        vpn_username = data.get('vpn_username', '')
        vpn_password = data.get('vpn_password', '')
        
        if not all([sender_email, vpn_username, vpn_password]):
            return jsonify({'error': 'Missing required parameters'}), 400
        
        # Execute automation
        from .automation_service import get_automation_service
        automation_service = get_automation_service()
        
        results = automation_service.execute_vpn_creation_automation(
            sender_email, vpn_username, vpn_password
        )
        
        return jsonify(results)
        
    except Exception as e:
        logging.error(f"Error executing VPN automation: {e}")
        return jsonify({'error': f'Automation failed: {str(e)}'}), 500


@main.route('/api/automation/vpn-creation/status', methods=['GET'])
@login_required
def get_automation_status():
    """Get status of ongoing automation"""
    # This could be enhanced with real-time status tracking
    return jsonify({
        'success': True,
        'status': 'ready',
        'message': 'Automation service is ready'
    })


@main.route('/api/ai/knowledge-status')
@login_required
def knowledge_base_status():
    """Check knowledge base loading status and test integration"""
    try:
        from .ai_service import get_ai_service
        ai_service = get_ai_service()
        
        status = ai_service.get_knowledge_base_status()
        
        # Test knowledge base with a sample question
        test_response = ai_service.test_knowledge_base_integration()
        
        return jsonify({
            'success': True,
            'knowledge_base': status,
            'test_question': "How do I reset a user's VPN access?",
            'test_response': test_response
        })
        
    except Exception as e:
        logging.error(f"Error checking knowledge base status: {e}")
        return jsonify({'error': f'Knowledge base status check failed: {str(e)}'}), 500


@main.route('/api/ai/get-knowledge-content')
@login_required
def get_knowledge_content():
    """Get current knowledge base content for editing"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Admin access required'}), 403
        
    try:
        from .ai_service import get_ai_service
        ai_service = get_ai_service()
        
        return jsonify({
            'success': True,
            'content': ai_service.knowledge_base or ''
        })
        
    except Exception as e:
        logging.error(f"Error getting knowledge content: {e}")
        return jsonify({'error': f'Failed to get knowledge content: {str(e)}'}), 500


@main.route('/api/ai/get-knowledge-status')
@login_required
def get_knowledge_status():
    """Get current knowledge base status and source information"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Admin access required'}), 403
        
    try:
        from .ai_service import get_ai_service
        ai_service = get_ai_service()
        
        status = ai_service.get_knowledge_base_status()
        return jsonify({
            'success': True,
            'status': status
        })
        
    except Exception as e:
        logging.error(f"Error getting knowledge status: {e}")
        return jsonify({'error': f'Failed to get knowledge status: {str(e)}'}), 500


@main.route('/main/update_knowledge_base', methods=['POST'])
@login_required
def update_knowledge_base():
    """Update the knowledge base content"""
    if current_user.role != 'admin':
        flash('Admin access required to update knowledge base.', 'error')
        return redirect(url_for('main.settings'))
        
    try:
        knowledge_content = request.form.get('knowledge_base_content', '').strip()
        
        if not knowledge_content:
            flash('Knowledge base content cannot be empty.', 'error')
            return redirect(url_for('main.settings'))
        
        # Get current status for audit log
        from .ai_service import get_ai_service
        ai_service = get_ai_service()
        old_status = ai_service.get_knowledge_base_status()
        old_content_length = old_status.get('content_length', 0)
        old_source_path = old_status.get('source_path', 'Unknown')
        
        # Update the AI service knowledge base using proper method
        success = ai_service.update_knowledge_base(knowledge_content)
        
        if success:
            # Get new status for audit log
            new_status = ai_service.get_knowledge_base_status()
            new_content_length = new_status.get('content_length', 0)
            new_source_path = new_status.get('source_path', 'Unknown')
            
            # Create audit log entry
            from .models import Log
            audit_entry = Log(
                user=current_user.username,
                action='Knowledge Base Updated',
                details=f'Updated knowledge base content. Size: {old_content_length} → {new_content_length} characters. Source: {old_source_path} → {new_source_path}'
            )
            db.session.add(audit_entry)
            db.session.commit()
            
            flash('Knowledge base updated successfully! The chatbot will now use the new content.', 'success')
            logging.info(f"Knowledge base updated by {current_user.username}: {old_content_length} → {new_content_length} characters")
        else:
            # Log failed attempt
            from .models import Log
            audit_entry = Log(
                user=current_user.username,
                action='Knowledge Base Update Failed',
                details=f'Failed to update knowledge base content. Content size: {len(knowledge_content)} characters'
            )
            db.session.add(audit_entry)
            db.session.commit()
            
            flash('Failed to update knowledge base. Please try again.', 'error')
        
    except Exception as e:
        logging.error(f"Error updating knowledge base: {e}")
        
        # Log error in audit
        try:
            from .models import Log
            audit_entry = Log(
                user=current_user.username,
                action='Knowledge Base Update Error',
                details=f'Error occurred while updating knowledge base: {str(e)}'
            )
            db.session.add(audit_entry)
            db.session.commit()
        except:
            pass  # Don't fail if audit logging fails
            
        flash(f'An error occurred while updating the knowledge base: {str(e)}', 'error')
    
    return redirect(url_for('main.settings'))


@main.route('/api/ai/reset-knowledge-base', methods=['POST'])
@login_required
@csrf.exempt
def reset_knowledge_base():
    """Reset knowledge base to original document"""
    logging.info("Reset knowledge base route accessed")
    
    if current_user.role != 'admin':
        logging.warning(f"Non-admin user {current_user.username} attempted to reset knowledge base")
        return jsonify({'error': 'Admin access required'}), 403
        
    try:
        logging.info("Attempting to reset knowledge base...")
        from .ai_service import get_ai_service
        ai_service = get_ai_service()
        
        # Get current status for audit log
        old_status = ai_service.get_knowledge_base_status()
        old_content_length = old_status.get('content_length', 0)
        old_source_path = old_status.get('source_path', 'Unknown')
        old_source_type = old_status.get('source_type', 'unknown')
        
        # Use the proper reset method from AI service
        success = ai_service.reset_knowledge_base()
        logging.info(f"Reset knowledge base result: {success}")
        
        if success:
            # Get new status for audit log
            new_status = ai_service.get_knowledge_base_status()
            new_content_length = new_status.get('content_length', 0)
            new_source_path = new_status.get('source_path', 'Unknown')
            new_source_type = new_status.get('source_type', 'unknown')
            
            # Create audit log entry
            from .models import Log
            audit_entry = Log(
                user=current_user.username,
                action='Knowledge Base Reset',
                details=f'Reset knowledge base to original document. Previous: {old_source_type} ({old_content_length} chars, {old_source_path}) → Current: {new_source_type} ({new_content_length} chars, {new_source_path})'
            )
            db.session.add(audit_entry)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Knowledge base reset to original document'
            })
        else:
            # Log failed attempt
            from .models import Log
            audit_entry = Log(
                user=current_user.username,
                action='Knowledge Base Reset Failed',
                details=f'Failed to reset knowledge base. Current state: {old_source_type} ({old_content_length} chars, {old_source_path})'
            )
            db.session.add(audit_entry)
            db.session.commit()
            
            return jsonify({
                'success': False,
                'error': 'Failed to reset knowledge base'
            }), 500
        
    except Exception as e:
        logging.error(f"Error resetting knowledge base: {e}")
        
        # Log error in audit
        try:
            from .models import Log
            audit_entry = Log(
                user=current_user.username,
                action='Knowledge Base Reset Error',
                details=f'Error occurred while resetting knowledge base: {str(e)}'
            )
            db.session.add(audit_entry)
            db.session.commit()
        except:
            pass  # Don't fail if audit logging fails
            
        return jsonify({'error': f'Failed to reset knowledge base: {str(e)}'}), 500


@main.route('/api/ai/chatbot', methods=['POST'])
@login_required
@csrf.exempt
def ai_chatbot():
    """AI-powered chatbot endpoint for user assistance."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'response': 'No data provided'
            }), 200
        
        message = data.get('message', '')
        ticket_id = data.get('ticket_id')
        
        if not message:
            return jsonify({
                'success': False,
                'response': 'Message is required'
            }), 200
        
        # Get ticket context if provided
        ticket_context = None
        if ticket_id:
            ticket = Ticket.query.get(ticket_id)
            if ticket:
                # Get assigned user if any
                assigned_user = None
                if ticket.assigned_to:
                    assigned_user = User.query.get(ticket.assigned_to)
                
                ticket_context = {
                    'id': ticket.id,
                    'subject': ticket.subject,
                    'body': ticket.description,  # Note: the field is 'description', not 'body'
                    'sender': ticket.sender,
                    'category': ticket.category,  # category is already a string
                    'status': ticket.status,
                    'urgency': ticket.urgency,
                    'created_at': ticket.created_at.strftime('%Y-%m-%d %H:%M') if ticket.created_at else 'Unknown',
                    'assigned_to': assigned_user.username if assigned_user else 'Unassigned',
                    'recent_activity': []  # Log system doesn't support ticket-specific logs yet
                }
        
        ai_service = get_ai_service()
        
        # Create user context for the chatbot
        user_context = {
            'username': current_user.username,
            'role': current_user.role,
            'id': current_user.id
        }
        
        response = ai_service.chatbot_response(message, ticket_context, user_context)
        
        return jsonify({
            'success': True,
            'response': response
        })
        
    except Exception as e:
        logging.error(f"Error in chatbot: {e}")
        return jsonify({
            'success': False,
            'response': f'Chatbot request failed: {str(e)}'
        }), 200


@main.route('/api/ai/analyze-sentiment', methods=['POST'])
@login_required
@csrf.exempt
def ai_analyze_sentiment():
    """AI-powered sentiment analysis endpoint."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        text = data.get('text', '')
        
        if not text:
            return jsonify({'error': 'Text is required'}), 400
        
        # Check if OpenAI API key is configured
        if not os.getenv('OPENAI_API_KEY'):
            return jsonify({'error': 'OpenAI API key not configured'}), 500
        
        ai_service = get_ai_service()
        result = ai_service.analyze_ticket_sentiment(text)
        
        return jsonify({
            'success': True,
            'sentiment': result
        })
        
    except Exception as e:
        logging.error(f"Error in sentiment analysis: {e}")
        return jsonify({'error': f'Sentiment analysis failed: {str(e)}'}), 500


@main.route('/api/ai/auto-categorize/<int:ticket_id>', methods=['POST'])
@login_required
@csrf.exempt
def ai_auto_categorize_ticket(ticket_id):
    """Automatically categorize an existing ticket using AI."""
    try:
        ticket = Ticket.query.get_or_404(ticket_id)
        
        # Check if OpenAI API key is configured
        if not os.getenv('OPENAI_API_KEY'):
            return jsonify({'error': 'OpenAI API key not configured'}), 500
        
        ai_service = get_ai_service()
        result = ai_service.categorize_ticket(ticket.subject, ticket.body, ticket.sender)
        
        # Find the category and update the ticket if confidence is high enough
        if result['confidence'] > 0.7:  # Only auto-update if confidence > 70%
            category = Category.query.filter_by(name=result['category_name']).first()
            if category:
                old_category = ticket.category.name if ticket.category else 'None'
                ticket.category_id = category.id
                ticket.urgency = result['urgency']
                db.session.commit()
                
                # Log the change
                log = TicketLog(
                    user=current_user.username,
                    action='ai_categorize',
                    details=f"AI categorized ticket '{ticket.subject}' (ID: {ticket.id}) from '{old_category}' to '{category.name}' with {result['confidence']:.1%} confidence. Urgency set to {result['urgency']}."
                )
                db.session.add(log)
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'updated': True,
                    'categorization': result,
                    'message': f'Ticket categorized as "{category.name}" with {result["confidence"]:.1%} confidence'
                })
        
        return jsonify({
            'success': True,
            'updated': False,
            'categorization': result,
            'message': f'AI suggests "{result["category_name"]}" category with {result["confidence"]:.1%} confidence (not auto-applied due to low confidence)'
        })
        
    except Exception as e:
        logging.error(f"Error in auto-categorization: {e}")
        return jsonify({'error': f'Auto-categorization failed: {str(e)}'}), 500


@main.route('/open_in_outlook/<int:ticket_id>')
@login_required
def open_in_outlook(ticket_id):
    """Open original email directly from email server in Outlook"""
    ticket = Ticket.query.get_or_404(ticket_id)
    
    # Role-based access control
    if current_user.role == 'infra':
        if ticket.assigned_to is not None and ticket.assigned_to != current_user.id:
            flash('You do not have permission to view this ticket.', 'error')
            return redirect(url_for('main.tickets'))
    
    # Get the latest email from thread if available
    from .models import EmailMessage
    latest_email = None
    if ticket.thread_id:
        latest_email = EmailMessage.query.filter_by(
            thread_id=ticket.thread_id
        ).order_by(EmailMessage.sent_at.desc()).first()
    
    # Get message ID for IMAP lookup
    message_id = None
    if latest_email and latest_email.message_id:
        message_id = latest_email.message_id
    
    if not message_id:
        # Fallback to EML approach if no message_id
        return download_eml_fallback(ticket, latest_email)
    
    try:
        # Connect to IMAP server to get original email
        import imaplib
        import os
        from dotenv import load_dotenv
        
        load_dotenv()
        
        # IMAP connection
        imap_server = os.getenv('IMAP_SERVER', 'imap.gmail.com')
        imap_port = int(os.getenv('IMAP_PORT', 993))
        email_user = os.getenv('GMAIL_USER')
        email_pass = os.getenv('GMAIL_APP_PASSWORD')
        
        if not email_user or not email_pass:
            flash('Email credentials not configured', 'error')
            return download_eml_fallback(ticket, latest_email)
        
        # Connect to IMAP
        mail = imaplib.IMAP4_SSL(imap_server, imap_port)
        mail.login(email_user, email_pass)
        mail.select('inbox')
        
        # Search for the email by Message-ID
        search_criteria = f'HEADER Message-ID "{message_id}"'
        status, messages = mail.search(None, search_criteria)
        
        if status != 'OK' or not messages[0]:
            mail.logout()
            flash('Original email not found on server', 'warning')
            return download_eml_fallback(ticket, latest_email)
        
        # Get the email
        email_id = messages[0].split()[-1]
        status, msg_data = mail.fetch(email_id, '(RFC822)')
        
        if status != 'OK':
            mail.logout()
            flash('Failed to fetch original email', 'error')
            return download_eml_fallback(ticket, latest_email)
        
        # Get the raw email content and parse it
        raw_email = msg_data[0][1]
        mail.logout()
        
        # Parse the email to modify headers for proper Outlook display
        import email as email_lib
        msg = email_lib.message_from_bytes(raw_email)
        
        # Preserve all original headers but ensure proper From/To perspective
        original_from = msg.get('From', '')
        original_to = msg.get('To', '')
        original_cc = msg.get('Cc', '')
        original_bcc = msg.get('Bcc', '')
        original_reply_to = msg.get('Reply-To', '')
        
        # Set tebstrack@gmail.com as the recipient (shows in From field when viewing)
        tebstrack_email = 'tebstrack@gmail.com'
        tebs_infra_email = 'tebs_infra@gmail.com'
        
        # Clean up any CSS artifacts from HTML processing
        def clean_email_content(part):
            if part.get_content_type() == 'text/html':
                content = part.get_payload(decode=True)
                if content:
                    try:
                        charset = part.get_content_charset() or 'utf-8'
                        html_content = content.decode(charset, errors='replace')
                        # Remove CSS artifacts like "P {margin-top:0;margin-bottom:0;}"
                        import re
                        html_content = re.sub(r'P\s*\{[^}]*\}', '', html_content, flags=re.IGNORECASE)
                        html_content = re.sub(r'[A-Z]\s*\{[^}]*\}', '', html_content, flags=re.IGNORECASE)
                        # Remove standalone CSS properties
                        html_content = re.sub(r'margin-[a-z]+:\s*\d+[^;]*;?', '', html_content, flags=re.IGNORECASE)
                        part.set_payload(html_content.encode(charset))
                    except Exception:
                        pass
        
        # Clean content in all parts
        if msg.is_multipart():
            for part in msg.walk():
                clean_email_content(part)
        else:
            clean_email_content(msg)
        
        # Replace headers to show proper perspective for Outlook viewing
        # When viewing in Outlook, we want to show the email as it was actually sent:
        # From: The actual sender (preserve original)
        # To: The original recipients (preserve unless tebstrack was the sender)
        
        tebstrack_emails = ['tebstrack@gmail.com', 'tebs_infra@totalebizsolutions.com']
        
        # Keep the original From as is (the actual sender)
        # Don't change it - that's correct now
        
        # Only clean the To field if tebstrack was the sender (in From field)
        # This means tebstrack sent an email, so we should remove tebstrack from To
        tebstrack_was_sender = any(teb_email.lower() in original_from.lower() for teb_email in tebstrack_emails)
        
        if tebstrack_was_sender and original_to:
            # tebstrack was the sender, so remove tebstrack emails from To field
            to_addresses = [addr.strip() for addr in original_to.split(',')]
            filtered_addresses = []
            for addr in to_addresses:
                # Check if this address contains any tebstrack email
                is_tebstrack = any(teb_email.lower() in addr.lower() for teb_email in tebstrack_emails)
                if not is_tebstrack:
                    filtered_addresses.append(addr)
            
            cleaned_to = ', '.join(filtered_addresses) if filtered_addresses else ''
            if cleaned_to:
                if 'To' in msg:
                    msg.replace_header('To', cleaned_to)
                else:
                    msg['To'] = cleaned_to
            else:
                # If no valid To addresses remain, remove the To header
                if 'To' in msg:
                    del msg['To']
        else:
            # tebstrack was the recipient, so preserve the To field as is
            # This is the normal case where someone sent TO tebstrack
            pass  # Keep original To field unchanged
        
        # Set Reply-To header to ensure replies go to tebstrack
        # Priority: tebstrack@gmail.com, then tebs_infra@totalebizsolutions.com
        reply_to_email = 'tebstrack@gmail.com'
        if original_to and 'tebs_infra@totalebizsolutions.com' in original_to.lower():
            reply_to_email = 'tebs_infra@totalebizsolutions.com'
        
        # Configure headers for proper "Reply All" behavior in Outlook
        # When user clicks "Reply All" in Outlook:
        # - To field should contain original sender + other non-tebstrack recipients
        # - CC field should contain original CC recipients (excluding tebstrack)
        # - tebstrack emails should never appear in To/CC of the reply
        
        # Set Reply-To to the appropriate tebstrack email for replies
        if 'Reply-To' in msg:
            msg.replace_header('Reply-To', reply_to_email)
        else:
            msg['Reply-To'] = reply_to_email
        
        # Add Return-Path for better email client handling
        if 'Return-Path' in msg:
            msg.replace_header('Return-Path', f'<{reply_to_email}>')
        else:
            msg['Return-Path'] = f'<{reply_to_email}>'
        
        # Set up headers to ensure Outlook's Reply All excludes tebstrack from To/CC
        # We need to manipulate the perspective so tebstrack appears as the "receiver"
        # not as a participant in the conversation
        
        # Clean CC field if it exists - remove tebstrack emails from CC
        if original_cc:
            cc_addresses = [addr.strip() for addr in original_cc.split(',')]
            filtered_cc = []
            for addr in cc_addresses:
                is_tebstrack = any(teb_email.lower() in addr.lower() for teb_email in ['tebstrack@gmail.com', 'tebs_infra@totalebizsolutions.com'])
                if not is_tebstrack:
                    filtered_cc.append(addr)
            
            cleaned_cc = ', '.join(filtered_cc) if filtered_cc else ''
            if cleaned_cc:
                if 'Cc' in msg:
                    msg.replace_header('Cc', cleaned_cc)
                else:
                    msg['Cc'] = cleaned_cc
            else:
                # Remove CC header if only tebstrack emails were there
                if 'Cc' in msg:
                    del msg['Cc']
        
        # Add custom headers to help email clients understand the reply structure
        msg['X-Original-Sender'] = original_from
        msg['X-TeBSTrack-Recipient'] = reply_to_email
        
        # Preserve BCC if they exist, but clean out tebstrack emails
        if original_bcc:
            bcc_addresses = [addr.strip() for addr in original_bcc.split(',')]
            filtered_bcc = []
            for addr in bcc_addresses:
                is_tebstrack = any(teb_email.lower() in addr.lower() for teb_email in ['tebstrack@gmail.com', 'tebs_infra@totalebizsolutions.com'])
                if not is_tebstrack:
                    filtered_bcc.append(addr)
            
            cleaned_bcc = ', '.join(filtered_bcc) if filtered_bcc else ''
            if cleaned_bcc:
                if 'Bcc' in msg:
                    msg.replace_header('Bcc', cleaned_bcc)
                else:
                    msg['Bcc'] = cleaned_bcc
            else:
                # Remove BCC header if only tebstrack emails were there
                if 'Bcc' in msg:
                    del msg['Bcc']
        
        # Convert back to bytes
        modified_email = msg.as_bytes()
        
        # Create response with modified email
        response = make_response(modified_email)
        response.headers['Content-Type'] = 'message/rfc822'
        response.headers['Content-Disposition'] = f'attachment; filename="original_ticket_{ticket.id}.eml"'
        
        return response
        
    except Exception as e:
        logging.error(f"Failed to fetch original email: {e}")
        flash('Could not access original email from server', 'error')
        return download_eml_fallback(ticket, latest_email)


def download_eml_fallback(ticket, latest_email=None):
    """Fallback EML creation when original email is not accessible"""
    # Use latest email data or fall back to ticket data
    if latest_email:
        subject = latest_email.subject
        body = latest_email.body
        sender = latest_email.sender
        date = latest_email.sent_at.strftime('%a, %d %b %Y %H:%M:%S %z') if latest_email.sent_at else 'Unknown'
        message_id = latest_email.message_id or f"<tebstrack-{ticket.id}@tebstrack>"
    else:
        subject = ticket.subject
        body = ticket.description
        sender = ticket.sender
        date = ticket.created_at.strftime('%a, %d %b %Y %H:%M:%S %z') if ticket.created_at else 'Unknown'
        message_id = f"<tebstrack-{ticket.id}@tebstrack>"
    
    # Clean body content from CSS artifacts
    import re
    if body:
        # Remove CSS artifacts like "P {margin-top:0;margin-bottom:0;}"
        body = re.sub(r'P\s*\{[^}]*\}', '', body, flags=re.IGNORECASE)
        body = re.sub(r'[A-Z]\s*\{[^}]*\}', '', body, flags=re.IGNORECASE)
        # Remove standalone CSS properties
        body = re.sub(r'margin-[a-z]+:\s*\d+[^;]*;?', '', body, flags=re.IGNORECASE)
        # Remove any stray CSS fragments
        body = re.sub(r'\{[^}]*\}', '', body)
        body = body.strip()
    
    # Create EML content with proper headers and reply configuration
    # From shows the actual sender, To shows original recipients (cleaned appropriately)
    # Headers configured for proper "Reply All" behavior in Outlook
    
    # Determine appropriate Reply-To address
    reply_to_email = 'tebstrack@gmail.com'
    if 'tebs_infra@totalebizsolutions.com' in str(latest_email.to_addresses if latest_email else ''):
        reply_to_email = 'tebs_infra@totalebizsolutions.com'
    
    # For fallback, we typically don't have the original recipient details
    # so we'll create a basic structure that ensures proper Reply All behavior
    to_field = ''
    if latest_email and hasattr(latest_email, 'to_addresses') and latest_email.to_addresses:
        # Clean tebstrack emails from To field for Reply All purposes
        to_addresses = [addr.strip() for addr in latest_email.to_addresses.split(',')]
        filtered_to = []
        for addr in to_addresses:
            is_tebstrack = any(teb_email.lower() in addr.lower() for teb_email in ['tebstrack@gmail.com', 'tebs_infra@totalebizsolutions.com'])
            if not is_tebstrack:
                filtered_to.append(addr)
        
        if filtered_to:
            to_field = f"To: {', '.join(filtered_to)}\n"
    
    eml_content = f"""From: {sender}
{to_field}Reply-To: {reply_to_email}
Return-Path: <{reply_to_email}>
X-Original-Sender: {sender}
X-TeBSTrack-Recipient: {reply_to_email}
Subject: {subject}
Date: {date}
Message-ID: {message_id}
MIME-Version: 1.0
Content-Type: text/html; charset=utf-8
Content-Transfer-Encoding: 8bit

<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{subject}</title>
    <style>
        body {{ font-family: Arial, sans-serif; font-size: 14px; line-height: 1.6; margin: 0; padding: 20px; }}
        .email-content {{ max-width: 800px; margin: 0 auto; }}
    </style>
</head>
<body>
    <div class="email-content">
        {body.replace(chr(10), '<br>') if body else ''}
    </div>
</body>
</html>
"""
    
    # Create response
    response = make_response(eml_content)
    response.headers['Content-Type'] = 'message/rfc822'
    response.headers['Content-Disposition'] = f'attachment; filename="ticket_{ticket.id}.eml"'
    
    return response


# ============================================================================
# EMAIL TEMPLATE MANAGEMENT - Admin Interface for Template Configuration
# ============================================================================

@main.route('/admin/email-templates')
@login_required
def manage_email_templates():
    """Admin interface for managing email templates"""
    if current_user.role != 'admin':
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('main.home'))
    
    from .models import EmailTemplate, TemplateActionStep
    
    templates = EmailTemplate.query.order_by(EmailTemplate.created_at.desc()).all()
    
    # Get action step counts for each template
    template_data = []
    for template in templates:
        step_count = TemplateActionStep.query.filter_by(template_id=template.id).count()
        template_data.append({
            'template': template,
            'step_count': step_count
        })
    
    return render_template('admin/email_templates.html', template_data=template_data)

@main.route('/admin/email-templates/new', methods=['GET', 'POST'])
@login_required
def create_email_template():
    """Create a new email template"""
    if current_user.role != 'admin':
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('main.home'))
    
    from .models import EmailTemplate
    
    if request.method == 'POST':
        try:
            template = EmailTemplate(
                name=request.form['name'],
                subject=request.form['subject'],
                body=request.form['body'],
                use_case_description=request.form.get('use_case_description', ''),
                is_active=request.form.get('is_active') == 'on',
                created_by=current_user.id
            )
            
            db.session.add(template)
            db.session.commit()
            
            # Add audit log
            log = TicketLog(
                user=current_user.username,
                action='template_create',
                details=f'Created email template: {template.name}'
            )
            db.session.add(log)
            db.session.commit()
            
            flash('Email template created successfully!', 'success')
            return redirect(url_for('main.manage_email_templates'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating template: {str(e)}', 'error')
    
    return render_template('admin/create_email_template.html')

@main.route('/admin/email-templates/<int:template_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_email_template(template_id):
    """Edit an email template"""
    if current_user.role != 'admin':
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('main.home'))
    
    from .models import EmailTemplate
    
    template = EmailTemplate.query.get_or_404(template_id)
    
    if request.method == 'POST':
        try:
            template.name = request.form['name']
            template.subject = request.form['subject']
            template.body = request.form['body']
            template.use_case_description = request.form.get('use_case_description', '')
            template.is_active = request.form.get('is_active') == 'on'
            
            db.session.commit()
            
            # Add audit log
            log = TicketLog(
                user=current_user.username,
                action='template_update',
                details=f'Updated email template: {template.name}'
            )
            db.session.add(log)
            db.session.commit()
            
            flash('Email template updated successfully!', 'success')
            return redirect(url_for('main.manage_email_templates'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating template: {str(e)}', 'error')
    
    return render_template('admin/edit_email_template.html', template=template)

@main.route('/admin/email-templates/<int:template_id>/delete', methods=['DELETE'])
@csrf.exempt
@login_required
def delete_email_template(template_id):
    """Delete an email template"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'You do not have permission to perform this action.'}), 403
    
    template = EmailTemplate.query.get_or_404(template_id)
    
    try:
        # Delete associated action steps first
        TemplateActionStep.query.filter_by(template_id=template_id).delete()
        
        # Delete template recommendations
        TicketTemplateRecommendation.query.filter_by(template_id=template_id).delete()
        
        # Store template name for logging before deletion
        template_name = template.name
        
        # Delete the template
        db.session.delete(template)
        db.session.commit()
        
        # Add audit log
        log = TicketLog(
            user=current_user.username,
            action='template_delete',
            details=f'Deleted email template: {template_name}'
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Template deleted successfully!'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error deleting template: {str(e)}'}), 500

@main.route('/admin/email-templates/<int:template_id>/action-steps')
@login_required
def manage_template_action_steps(template_id):
    """Manage action steps for an email template"""
    if current_user.role != 'admin':
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('main.home'))
    
    from .models import EmailTemplate, TemplateActionStep
    
    template = EmailTemplate.query.get_or_404(template_id)
    action_steps = TemplateActionStep.query.filter_by(template_id=template_id).order_by(TemplateActionStep.step_order).all()
    
    return render_template('admin/template_action_steps.html', template=template, action_steps=action_steps)

@main.route('/admin/email-templates/<int:template_id>/action-steps/new', methods=['GET', 'POST'])
@login_required
def create_action_step(template_id):
    """Create a new action step for a template"""
    if current_user.role != 'admin':
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('main.home'))
    
    from .models import EmailTemplate, TemplateActionStep
    
    template = EmailTemplate.query.get_or_404(template_id)
    
    if request.method == 'POST':
        try:
            # Get the next order number
            max_order = db.session.query(db.func.max(TemplateActionStep.step_order)).filter_by(template_id=template_id).scalar() or 0
            
            action_step = TemplateActionStep(
                template_id=template_id,
                step_order=max_order + 1,
                step_type=request.form['step_type'],
                step_title=request.form['step_title'],
                step_description=request.form['step_description'],
                is_automated=request.form.get('is_automated') == 'on',
                step_config=request.form.get('step_config', '{}') if request.form.get('step_config') else '{}'
            )
            
            db.session.add(action_step)
            db.session.commit()
            
            # Add audit log
            log = TicketLog(
                user=current_user.username,
                action='action_step_create',
                details=f'Created action step for template: {template.name}'
            )
            db.session.add(log)
            db.session.commit()
            
            flash('Action step created successfully!', 'success')
            return redirect(url_for('main.manage_template_action_steps', template_id=template_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating action step: {str(e)}', 'error')
    
    return render_template('admin/create_action_step.html', template=template)

@main.route('/admin/action-steps/<int:step_id>/edit')
@login_required
def edit_action_step(step_id):
    """Edit an action step"""
    if current_user.role != 'admin':
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('main.home'))
    
    step = TemplateActionStep.query.get_or_404(step_id)
    template = EmailTemplate.query.get_or_404(step.template_id)
    
    if request.method == 'POST':
        try:
            step.step_title = request.form['step_title']
            step.step_type = request.form['step_type'] 
            step.step_description = request.form['step_description']
            step.is_automated = request.form.get('is_automated') == 'on'
            step.step_config = request.form.get('step_config', '{}') if request.form.get('step_config') else '{}'
            
            db.session.commit()
            
            # Add audit log
            log = TicketLog(
                user=current_user.username,
                action='action_step_edit',
                details=f'Edited action step: {step.step_title} for template: {template.name}'
            )
            db.session.add(log)
            db.session.commit()
            
            flash('Action step updated successfully!', 'success')
            return redirect(url_for('main.manage_template_action_steps', template_id=template.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating action step: {str(e)}', 'error')
    
    return render_template('admin/edit_action_step.html', step=step, template=template)

@main.route('/admin/action-steps/<int:step_id>/move', methods=['POST'])
@csrf.exempt
@login_required
def move_action_step(step_id):
    """Move an action step up or down in order"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'You do not have permission to perform this action.'}), 403
    
    try:
        data = request.get_json()
        direction = data.get('direction')
        
        step = TemplateActionStep.query.get_or_404(step_id)
        template = EmailTemplate.query.get_or_404(step.template_id)
        
        # Get all steps for this template ordered by step_order
        all_steps = TemplateActionStep.query.filter_by(template_id=step.template_id).order_by(TemplateActionStep.step_order).all()
        
        current_index = None
        for i, s in enumerate(all_steps):
            if s.id == step_id:
                current_index = i
                break
        
        if current_index is None:
            return jsonify({'success': False, 'message': 'Step not found'}), 404
        
        # Determine new position
        if direction == 'up' and current_index > 0:
            # Swap with previous step
            other_step = all_steps[current_index - 1]
            step.step_order, other_step.step_order = other_step.step_order, step.step_order
        elif direction == 'down' and current_index < len(all_steps) - 1:
            # Swap with next step
            other_step = all_steps[current_index + 1]
            step.step_order, other_step.step_order = other_step.step_order, step.step_order
        else:
            return jsonify({'success': False, 'message': f'Cannot move step {direction}'}), 400
        
        db.session.commit()
        
        # Add audit log
        log = TicketLog(
            user=current_user.username,
            action='action_step_move',
            details=f'Moved action step: {step.step_title} {direction} for template: {template.name}'
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Step moved {direction} successfully!'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error moving step: {str(e)}'}), 500

@main.route('/admin/action-steps/<int:step_id>/delete', methods=['DELETE'])
@csrf.exempt
@login_required
def delete_action_step(step_id):
    """Delete an action step"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'You do not have permission to perform this action.'}), 403
    
    try:
        step = TemplateActionStep.query.get_or_404(step_id)
        template = EmailTemplate.query.get_or_404(step.template_id)
        step_title = step.step_title
        
        # Get all steps with higher order numbers to reorder them
        higher_steps = TemplateActionStep.query.filter(
            TemplateActionStep.template_id == step.template_id,
            TemplateActionStep.step_order > step.step_order
        ).all()
        
        # Delete the step
        db.session.delete(step)
        
        # Reorder remaining steps
        for higher_step in higher_steps:
            higher_step.step_order -= 1
        
        db.session.commit()
        
        # Add audit log
        log = TicketLog(
            user=current_user.username,
            action='action_step_delete',
            details=f'Deleted action step: {step_title} from template: {template.name}'
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Action step deleted successfully!'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error deleting step: {str(e)}'}), 500

@main.route('/api/templates/recommend', methods=['POST'])
@csrf.exempt
@login_required
def recommend_template():
    """API endpoint for getting AI template recommendations"""
    try:
        data = request.get_json()
        
        if not data or 'subject' not in data or 'description' not in data:
            return jsonify({'error': 'Subject and description are required'}), 400
        
        from .ai_service import get_ai_service
        ai_service = get_ai_service()
        
        if not ai_service:
            return jsonify({'error': 'AI service not available'}), 503
        
        recommendation = ai_service.recommend_email_template(
            ticket_subject=data['subject'],
            ticket_description=data['description'],
            ticket_category=data.get('category')
        )
        
        return jsonify(recommendation)
        
    except Exception as e:
        logging.error(f"Template recommendation error: {e}")
        return jsonify({'error': f'Recommendation failed: {str(e)}'}), 500

@main.route('/api/templates/<template_name>/action-steps', methods=['POST'])
@csrf.exempt
@login_required
def get_template_action_steps(template_name):
    """API endpoint for getting action steps for a template"""
    try:
        data = request.get_json()
        
        ticket_context = {
            'subject': data.get('subject', ''),
            'description': data.get('description', ''),
            'category': data.get('category', ''),
            'sender': data.get('sender', ''),
            'id': data.get('ticket_id')
        }
        
        from .ai_service import get_ai_service
        ai_service = get_ai_service()
        
        if not ai_service:
            return jsonify({'error': 'AI service not available'}), 503
        
        action_steps = ai_service.generate_template_action_steps(template_name, ticket_context)
        
        return jsonify({'action_steps': action_steps})
        
    except Exception as e:
        logging.error(f"Action steps error: {e}")
        return jsonify({'error': f'Action steps generation failed: {str(e)}'}), 500
